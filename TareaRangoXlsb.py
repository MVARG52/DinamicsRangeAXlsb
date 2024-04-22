
import numpy as np
import csv

from pyxlsb import open_workbook
from openpyxl import load_workbook


hojaDinamic=1
valor=None

todos=False
parametros=['a','b',1,10]
encabezado=['campo1','campo2','campo3','campo4','campo5']

archivo_excel = 'C:/Users/USER/Desktop/Nueva carpeta/Practicas/Excels/par_convert_file.xlsx'
nombre_archivito = 'datos.csv'
def letra_a_numero(letra):
    valor_unicode = ord(letra.lower()) - ord('a')
    if 0 <= valor_unicode <= 25:
        return valor_unicode
    else:
        print("La entrada no es una letra del alfabeto.")
        return None

def convertExcel(archivo_excel,parametros,hojaDinamic,todos,valor,encabezado):
    columnaD= parametros[0]
    columnaH= parametros[1]
    filaD= parametros[2]
    filaH= parametros[3]
    try:
        #xlsb
        with open_workbook(archivo_excel.upper()) as libro:
            a=0
            b=0
            inicialV=0
            hoja = libro.get_sheet(hojaDinamic)       
            datos=[]
            if(todos==True):
                filaH=len(hoja)
                filaD=0
                columnaH=len(hoja[0])
                columnaD=0
            filas=filaH-filaD
            columnas=letra_a_numero(columnaH)-letra_a_numero(columnaD)
            for data in hoja.rows():
                datos.append(data)
            Len=len(datos)
            if(valor==None):
                Resultado=np.empty((filas+1, columnas+1), dtype=object)
                
            if(valor!=None):
                Resultado=np.empty((filas+2, columnas+2), dtype=object)
                for i in range(0,min(columnas, len(encabezado))):
                    Resultado[0][i]=encabezado[i]
                inicialV=1
            b=0
            for i in range(int(letra_a_numero(columnaD)), min(int(letra_a_numero(columnaH)) + 1, Len)):
                a=inicialV
                for j in range(filaD-1, min(filaH , Len)):
                    if(datos[j][i].v!=None):
                        Resultado[a][b]=datos[j][i].v
                    if(datos[j][i].v==None):
                        Resultado[a][b]=None
                    a=a+1
                b=b+1
        print('es binario')



    except:
    # Lee el archivo Excel protegido con contraseña
    
        libro = load_workbook(filename=archivo_excel)
        # Procesa el DataFrame
        
        hoja = libro.worksheets[hojaDinamic-1]
        a=0
        b=0     
        datosI=[]   
        dataTransi=[]
        inicialV=0
        for data in hoja.iter_rows():
            datosI.append(data)
        
        if(todos==True):
            filaH=len(datosI)
            filaD=0
            columnaH=len(datosI[0])
            columnaD=0
            columnas=len(datosI[0])
        if(todos==False):
            
            columnas=letra_a_numero(columnaH)-letra_a_numero(columnaD)
        filas=filaH-filaD
        b=0

        if(valor==None):
            Resultado=np.empty((filas+1, columnas+1), dtype=object)
            
        if(valor!=None):
            Resultado=np.empty((filas+2, columnas+2), dtype=object)
            for i in range(0,min(columnas, len(encabezado))):
                Resultado[0][i]=encabezado[i]
            inicialV=1
        if(todos==False):
            for i in range(int(letra_a_numero(columnaD)), min(int(letra_a_numero(columnaH))+1, int(letra_a_numero(columnaH))+1)):
                a=inicialV
                for j in range(filaD-1, min(filaH, filaH)):
                    Resultado[a][b]=datosI[j][i].value
                    a=a+1
                b=b+1
        if(todos==True): 
            for i in range(0, columnaH):
                a=0
                for j in range(0, min(filaH, filaH)):
                    Resultado[a][b]=datosI[j][i].value
                    a=a+1
                b=b+1

        print('No es binario')

def quitar_columnas(Resultado):
    #Quitar columnas nulas
    c=0
    a=0
    b=0
    ResultadoSinCol=[]
    ResultadoSinColR=np.empty((len(Resultado), len(Resultado[0])), dtype=object)
    for j in range(0, len(Resultado[0])):
        if(j!=0):
            if(c<len(Resultado)):
                b=b+1
        c=0
        
        for i in range(0, len(Resultado)):
            
            if(Resultado[i][j]==None):
                c=c+1
            
        for i in range(0, len(Resultado)):
            
            if(c<len(Resultado)):
                ResultadoSinColR[i][b]=Resultado[i][j]

    NumCol=b+1
    ResultadoSinColumnas=np.empty((len(ResultadoSinColR), NumCol), dtype=object)
    return(ResultadoSinColumnas)

def quitar_filas(ResultadoSinColR):
    ResultadoSinColumnas=np.empty((len(ResultadoSinColR), NumCol), dtype=object)
    NumCol=ResultadoSinColumnas[0]
    ##Quitar filas nulas
    dataTransi=[]
    for j in range(0, len(ResultadoSinColR)):
        b=0
        for Cols in range(0, NumCol):
            if(ResultadoSinColR[j][Cols]==None):
                b=b+1
        if(b<NumCol):
            dataTransi.append(ResultadoSinColR[j])
    b=0
    Resultado=np.empty((len(dataTransi), NumCol), dtype=object)
    for i in range(0, len(Resultado[0])):
        a=0
        for j in range(0, len(Resultado)):
            Resultado[a][i]=dataTransi[j][i]
            a=a+1
        b=b+1

    ResultadoSinFilas=np.empty((len(Resultado), b), dtype=object)

    return(ResultadoSinFilas)

def BarrerFilas(Resultado,inicialV):
    #print(ResultadoSinColR)
    #Barrer por filas
    print(Resultado)
    ResultadoConBarrido=Resultado
    for j in range(0, len(ResultadoConBarrido[0])):
        c=0
        fila=0
        lastValue=True
        lRow=inicialV
        for i in range(inicialV, len(ResultadoConBarrido)):
            if(Resultado[i][j]!=None):
                if(inicialV!=0):
                    ResultadoConBarrido[0][j]=Resultado[0][j]
                if isinstance(Resultado[i][j], (int,float,set)):
                    
                    ResultadoConBarrido[i][j]=Resultado[i][j]
                    lastValue=False
                if not isinstance(Resultado[i][j], (int,float,set)):
                    ResultadoConBarrido[i][j]=Resultado[i][j]
                    lRow=i
                    lastValue=True
            if(Resultado[i][j]==None):
                if(lastValue==True):
                    if(i!=0):
                        l=i
                        while(Resultado[l][j]==None):
                            ResultadoConBarrido[fila][j]=Resultado[lRow][j]
                            if(l<len(Resultado)-1):
                                l=l+1
                            else:
                                break    
            fila=fila+1
    return(ResultadoConBarrido)

def ToCsv(ResultadoConBarrido,archivoCSVname):
    with open(archivoCSVname, 'w', newline='') as archivo:
        escritor_csv = csv.writer(archivo, delimiter='|')
        for fila in ResultadoConBarrido:
            escritor_csv.writerow(fila)